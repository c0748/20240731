import openpyxl
from datetime import datetime
import os
from openpyxl.drawing.image import Image
import win32com.client as win32

# 請求書ダミーデータを読み込み
wb = openpyxl.load_workbook("files/invoice_data.xlsx", data_only=True)
ws = wb.active
values = list(ws.values)
lastrow = len(values)
print(lastrow)

# 2請求書のテンプレを読み込む
wb = openpyxl.load_workbook("files/invoice.xlsx", data_only=True)
ws = wb.active

# ３日付の取得とフォルダの作成

current_date = datetime.now()

invoice_date = current_date.strftime("%Y年%m月%d日")

year_month = current_date.strftime("%Y%m")

invoice_month = current_date.month

output_folder = f"請求書_{current_date.strftime('%Y年%m月')}"

os.makedirs(output_folder, exist_ok=True)

output_file = f"{output_folder}/請求書_{current_date.strftime('%Y年%m月')}.xlsx"

invoice_number = 1

successful_companies = []
failed_companies = []
no_invoice_companies = []


# 4繰り返し処理で請求書の中身を作成する
for index in range(lastrow):
    if not index == 0:
        if values[index][12] is None:
            no_invoice_companies.append(values[index][0])
            continue
        try:
            sheet_name = str(values[index][0])

            copy_ws = wb.copy_worksheet(ws)
            copy_ws.title = sheet_name
            # 会社名
            copy_ws["A2"].value = sheet_name
            # 担当者名
            copy_ws["A4"].value = values[index][10]
            # 件名
            copy_ws["B7"].value = f"{invoice_month}月分請求書"
            # 請求書番号
            copy_ws["N2"].value = f"{year_month}-{invoice_number:03d}"
            invoice_number += 1
            copy_ws["N3"].value = invoice_date

            # 摘要欄
            copy_ws["A14"].value = values[index][13]
            copy_ws["A15"].value = values[index][18]
            copy_ws["A16"].value = values[index][23]
            copy_ws["A17"].value = values[index][28]
            copy_ws["A18"].value = values[index][33]
            copy_ws["A19"].value = values[index][38]
            copy_ws["A20"].value = values[index][43]
            copy_ws["A21"].value = values[index][48]
            copy_ws["A22"].value = values[index][53]
            copy_ws["A23"].value = values[index][58]

            # 数量
            copy_ws["J14"].value = values[index][14]
            copy_ws["J15"].value = values[index][19]
            copy_ws["J16"].value = values[index][24]
            copy_ws["J17"].value = values[index][29]
            copy_ws["J18"].value = values[index][34]
            copy_ws["J19"].value = values[index][39]
            copy_ws["J20"].value = values[index][44]
            copy_ws["J21"].value = values[index][49]
            copy_ws["J22"].value = values[index][54]
            copy_ws["J23"].value = values[index][59]

            # 数量
            copy_ws["K14"].value = values[index][15]
            copy_ws["K15"].value = values[index][20]
            copy_ws["K16"].value = values[index][25]
            copy_ws["K17"].value = values[index][30]
            copy_ws["K18"].value = values[index][35]
            copy_ws["K19"].value = values[index][40]
            copy_ws["K20"].value = values[index][45]
            copy_ws["K21"].value = values[index][50]
            copy_ws["K22"].value = values[index][55]
            copy_ws["K23"].value = values[index][60]

            # 単価
            copy_ws["L14"].value = values[index][16]
            copy_ws["L15"].value = values[index][21]
            copy_ws["L16"].value = values[index][26]
            copy_ws["L17"].value = values[index][31]
            copy_ws["L18"].value = values[index][36]
            copy_ws["L19"].value = values[index][41]
            copy_ws["L20"].value = values[index][46]
            copy_ws["L21"].value = values[index][51]
            copy_ws["L22"].value = values[index][56]
            copy_ws["L23"].value = values[index][61]

            # 金額
            copy_ws["O14"].value = values[index][17]
            copy_ws["O15"].value = values[index][22]
            copy_ws["O16"].value = values[index][27]
            copy_ws["O17"].value = values[index][32]
            copy_ws["O18"].value = values[index][37]
            copy_ws["O19"].value = values[index][42]
            copy_ws["O20"].value = values[index][47]
            copy_ws["O21"].value = values[index][52]
            copy_ws["O22"].value = values[index][57]
            copy_ws["O23"].value = values[index][62]
            # 角印データ貼り付け

            img = Image("files/角印.png")

            img.width = 100

            img.height = 100

            copy_ws.add_image(img, "P5")

            successful_companies.append(values[index][0])
        except Exception as e:
            failed_companies.append(values[index][0])
        
wb.save(output_file)


def excel_to_pdf(sheet_name, output_pdf):
    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False

    wb = excel.Workbooks.Open(os.path.abspath(output_file))
    ws = wb.Sheets(sheet_name)

    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesWide = 1

    ws.ExportAsFixedFormat(0, os.path.abspath(output_pdf))
    wb.Close(False)
    excel.Application.Quit()


for sheet in wb.sheetnames:
    pdf_file = f"{output_folder}/{sheet}.pdf"
    excel_to_pdf(sheet, pdf_file)

current_date = datetime.now()
detail_log_file = f"{output_folder}/請求書発行報告_{current_date.strftime('%Y年%m月%d日')}.txt"
with open(detail_log_file, 'w', encoding='utf-8') as f:
    f.write(f"請求書発行日: {current_date.strftime('%Y年%m月%d日')}\n\n")
    f.write("1. 請求書発行成功会社:\n")
    for company in successful_companies:
        f.write(f"  - {company}\n")
    f.write(f"  合計: {len(successful_companies)}社\n\n")
    f.write("2. 請求書発行失敗会社:\n")
    for company in failed_companies:
        f.write(f"  - {company}\n")
    f.write(f"  合計: {len(failed_companies)}社\n\n")
    f.write("3. 請求書未発行会社:\n")
    for company in no_invoice_companies:
        f.write(f"  - {company}\n")
    f.write(f"  合計: {len(no_invoice_companies)}社\n\n")
    f.write(f"総登録会社数: {len(successful_companies) + len(failed_companies) + len(no_invoice_companies)}社")







