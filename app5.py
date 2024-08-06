import openpyxl
from datetime import datetime
import os
from openpyxl.drawing.image import Image
import win32com.client as win32


def load_invoice_data(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    return list(ws.values)

def create_invoice_template(file_path):
    return openpyxl.load_workbook(file_path)

def get_current_date_info():
    current_date = datetime.now()
    invoice_date = current_date.strftime("%Y年%-m月%d日")
    year_month = current_date.strftime("%Y%m")
    invoice_month = current_date.month
    output_folder = f"請求書_{current_date.strftime('%Y年%m月')}"
    os.makedirs(output_folder, exist_ok=True)
    output_file = f"{output_folder}/請求書_{current_date.strftime('%Y年%m月')}.xlsx"
    return invoice_date, year_month, invoice_month, output_folder, output_file

def fill_invoice_template(copy_ws, data, invoice_date, invoice_number, year_month, invoice_month):

    sheet_name = str(data[0])

    copy_ws.title = sheet_name

    copy_ws["A2"].value = sheet_name

    copy_ws["A4"].value = data[10]

    copy_ws["B7"].value = f"{invoice_month}月分請求書"

    copy_ws["N2"].value = f"{year_month}-{invoice_number:03d}"

    copy_ws["N3"].value = invoice_date

    

    for i in range(14, 24):

         copy_ws[f"A{i}"].value = data[13 + (i - 14) * 5]

         copy_ws[f"J{i}"].value = data[14 + (i - 14) * 5]

         copy_ws[f"K{i}"].value = data[15 + (i - 14) * 5]

         copy_ws[f"L{i}"].value = data[16 + (i - 14) * 5]

         copy_ws[f"O{i}"].value = data[17 + (i - 14) * 5]



    img = Image("files/角印.png")

    img.width = 100

    img.height = 100

    copy_ws.add_image(img, "P5")





# ステップ５ 請求書の作成を関数化する。
def generate_invoices(invoice_data, template_path, output_file):

    lastrow = len(invoice_data)

    wb = create_invoice_template(template_path)

    ws = wb.active

    invoice_date, year_month, invoice_month, output_folder, output_file = get_current_date_info()

    invoice_number = 1



    for index in range(lastrow):

            # 1行目はヘッダー部分なので取得しない

            if not index == 0:

                data = invoice_data[index]

                if data[12] is None:

                    continue



                copy_ws = wb.copy_worksheet(ws)

                fill_invoice_template(copy_ws, data, invoice_date, invoice_number, year_month, invoice_month)

                invoice_number += 1

    

    ws = wb["請求書"]

    wb.remove(ws)

    wb.save(output_file)

    return wb, output_folder, output_file


# ステップ６ PDF化を関数化。
def excel_to_pdf(sheet_name, output_pdf):

    excel = win32.gencache.EnsureDispatch('Excel.Application')

    excel.Visible = False

    

    wb = excel.Workbooks.Open(os.path.abspath(output_file))

    ws = wb.Sheets(sheet_name)

    ws.PageSetup.Zoom = False

    ws.PageSetup.FitToPagesWide = 1

    

    ws.ExportAsFixedFormat(0, os.path.abspath(output_pdf))

    wb.Close(False)

    excel.Application.Quit()



def generate_pdfs_from_excel(wb, output_folder):

    for sheet in wb.sheetnames:

        pdf_file = f"{output_folder}/{sheet}.pdf"

        excel_to_pdf(sheet, pdf_file)



# プログラムの順番をコントロール

def main():

    invoice_data = load_invoice_data("files/invoice_data.xlsx")

    template_path = "files/invoice.xlsx"

    wb, output_folder, output_file = generate_invoices(invoice_data, template_path, output_file)

    generate_pdfs_from_excel(wb, output_folder)



if __name__ == "__main__":

    main()
